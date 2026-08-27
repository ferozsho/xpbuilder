# Licensed to the Apache Software Foundation (ASF) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The ASF licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.
"""Report Designer — pages and JSON API.

Builds Crystal-Reports-style multi-table reports on top of existing Superset
datasets. Reports are persisted as JSON definitions and executed through the
underlying dataset databases.
"""

from __future__ import annotations

import csv
import logging
import re
from datetime import datetime
from io import StringIO
from typing import Any

from flask import g, request, Response
from flask_appbuilder import expose
from flask_appbuilder.security.decorators import has_access, has_access_api
from flask_babel import gettext as _

from superset import db
from superset.models.core import Database
from superset.models.report_designer import ReportDesigner
from superset.superset_typing import FlaskResponse
from superset.utils import core as utils
from superset.views.base import BaseSupersetView
from superset.views.error_handling import json_error_response

from .publish import (
    publish_report,
    PublishError,
    unpublish_report,
    VIZ_CATALOG,
)
from .sql_builder import (
    datasets_payload,
    execute_report,
    ReportDesignerError,
)
from .table_sync import (
    find_reporting_database,
    sync_tables,
    TableSyncError,
)

logger = logging.getLogger(__name__)

SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9_.-]+")


def _parse_definition() -> dict[str, Any]:
    """Parse the JSON definition from the request body."""
    payload = request.get_json(force=True, silent=True) or {}
    definition = payload.get("definition") if isinstance(payload, dict) else payload
    if not isinstance(definition, dict):
        raise ReportDesignerError(_("Report definition must be a JSON object"))
    return definition


def _save_report(
    report: ReportDesigner,
    name: str,
    description: str,
    definition: dict[str, Any],
    is_new: bool,
) -> ReportDesigner:
    """Persist a report definition with audit fields."""
    report.name = (name or "Untitled report").strip()
    report.description = description or ""
    report.set_definition(definition)
    now = datetime.now()
    user_id = utils.get_user_id()
    if is_new:
        report.created_on = now
        report.created_by_fk = user_id
    report.changed_on = now
    report.changed_by_fk = user_id
    db.session.add(report)
    db.session.commit()
    return report


class ReportDesignerView(BaseSupersetView):
    route_base = "/reportdesigner"
    class_permission_name = "ReportDesigner"

    # --- Pages ---------------------------------------------------------

    @expose("/list/")
    @has_access
    def list(self) -> FlaskResponse:
        return super().render_app_template()

    @expose("/designer/")
    @has_access
    def designer(self) -> FlaskResponse:
        return super().render_app_template()

    @expose("/designer/<int:report_id>/")
    @has_access
    def designer_detail(self, report_id: int) -> FlaskResponse:
        return super().render_app_template()

    # --- API -----------------------------------------------------------

    @expose("/api/", methods=("GET",))
    @has_access_api
    def api_list(self) -> FlaskResponse:
        reports = (
            db.session.query(ReportDesigner)
            .order_by(ReportDesigner.changed_on.desc())
            .limit(200)
            .all()
        )
        return self.json_response(
            {
                "count": len(reports),
                "result": [report.to_dict() for report in reports],
            }
        )

    @expose("/api/datasets/", methods=("GET",))
    @has_access_api
    def api_datasets(self) -> FlaskResponse:
        return self.json_response({"result": datasets_payload()})

    @expose("/api/sync-tables/", methods=("POST",))
    @has_access_api
    def api_sync_tables(self) -> FlaskResponse:
        """Register every table of the Moodle (read-only) database as a dataset."""
        try:
            payload = request.get_json(force=True, silent=True) or {}
            database_id = int(payload.get("database_id") or 0)
            if not database_id:
                database = find_reporting_database()
                database_id = database.id if database else 0
            if not database_id:
                return json_error_response(
                    _("No reporting database configured"), 400
                )
            return self.json_response(sync_tables(database_id))
        except TableSyncError as ex:
            return json_error_response(str(ex), 400)
        except Exception as ex:  # pylint: disable=broad-except
            db.session.rollback()
            logger.exception("Failed to sync tables")
            return json_error_response(utils.error_msg_from_exception(ex), 400)

    @expose("/api/viz-types/", methods=("GET",))
    @has_access_api
    def api_viz_types(self) -> FlaskResponse:
        """List the visualization types a report can be published as."""
        return self.json_response(
            {
                "result": [
                    {
                        "key": key,
                        "label": str(item["label"]),
                        "requires_dttm": bool(item.get("requires_dttm")),
                    }
                    for key, item in VIZ_CATALOG.items()
                ]
            }
        )

    @expose("/api/dashboards/", methods=("GET",))
    @has_access_api
    def api_dashboards(self) -> FlaskResponse:
        """List published Superset dashboards for the publish picker."""
        from superset.models.dashboard import Dashboard

        dashboards = (
            db.session.query(Dashboard.id, Dashboard.dashboard_title)
            .filter(Dashboard.published.is_(True))
            .order_by(Dashboard.dashboard_title.asc())
            .all()
        )
        return self.json_response(
            {
                "result": [
                    {"id": dashboard.id, "dashboard_title": dashboard.dashboard_title}
                    for dashboard in dashboards
                ]
            }
        )

    @expose("/api/", methods=("POST",))
    @has_access_api
    def api_post(self) -> FlaskResponse:
        try:
            payload = request.get_json(force=True, silent=True) or {}
            definition = payload.get("definition") or {}
            if not isinstance(definition, dict):
                return json_error_response(
                    _("Report definition must be a JSON object"), 400
                )
            report = _save_report(
                ReportDesigner(),
                payload.get("name", ""),
                payload.get("description", ""),
                definition,
                is_new=True,
            )
            return self.json_response(report.to_dict(), 201)
        except ReportDesignerError as ex:
            return json_error_response(str(ex), 400)
        except Exception as ex:  # pylint: disable=broad-except
            db.session.rollback()
            logger.exception("Failed to create report")
            return json_error_response(utils.error_msg_from_exception(ex), 400)

    @expose("/api/<int:report_id>/", methods=("GET",))
    @has_access_api
    def api_get(self, report_id: int) -> FlaskResponse:
        report = db.session.get(ReportDesigner, report_id)
        if report is None:
            return Response(status=404)
        return self.json_response(report.to_dict())

    @expose("/api/<int:report_id>/", methods=("PUT",))
    @has_access_api
    def api_put(self, report_id: int) -> FlaskResponse:
        try:
            report = db.session.get(ReportDesigner, report_id)
            if report is None:
                return Response(status=404)
            payload = request.get_json(force=True, silent=True) or {}
            definition = payload.get("definition") or report.get_definition()
            if not isinstance(definition, dict):
                return json_error_response(
                    _("Report definition must be a JSON object"), 400
                )
            _save_report(
                report,
                payload.get("name", report.name),
                payload.get("description", report.description or ""),
                definition,
                is_new=False,
            )
            return self.json_response(report.to_dict())
        except ReportDesignerError as ex:
            return json_error_response(str(ex), 400)
        except Exception as ex:  # pylint: disable=broad-except
            db.session.rollback()
            logger.exception("Failed to update report")
            return json_error_response(utils.error_msg_from_exception(ex), 400)

    @expose("/api/<int:report_id>/", methods=("DELETE",))
    @has_access_api
    def api_delete(self, report_id: int) -> FlaskResponse:
        report = db.session.get(ReportDesigner, report_id)
        if report is None:
            return Response(status=404)
        db.session.delete(report)
        db.session.commit()
        # Return 200 with a JSON body: SupersetClient's default parse method
        # is JSON, and a 204 No Content (empty body) makes response.json()
        # reject with "Unexpected end of JSON input".
        return self.json_response({"message": _("Report deleted")})

    @expose("/api/preview/", methods=("POST",))
    @has_access_api
    def api_preview(self) -> FlaskResponse:
        try:
            definition = _parse_definition()
            result = execute_report(definition)
            return self.json_response(result)
        except ReportDesignerError as ex:
            return json_error_response(str(ex), 400)
        except Exception as ex:  # pylint: disable=broad-except
            logger.exception("Failed to preview report")
            return json_error_response(utils.error_msg_from_exception(ex), 400)

    @expose("/api/<int:report_id>/publish/", methods=("POST",))
    @has_access_api
    def api_publish(self, report_id: int) -> FlaskResponse:
        """Publish a report as a Superset chart attached to a dashboard."""
        try:
            report = db.session.get(ReportDesigner, report_id)
            if report is None:
                return Response(status=404)
            payload = request.get_json(force=True, silent=True) or {}
            user = getattr(g, "user", None)
            if user is None or getattr(user, "is_anonymous", False):
                user = None
            result = publish_report(
                report,
                viz_key=str(payload.get("viz_type") or "table"),
                chart_name=payload.get("chart_name"),
                dashboard_id=(
                    int(payload["dashboard_id"])
                    if payload.get("dashboard_id")
                    else None
                ),
                new_dashboard_name=payload.get("new_dashboard_name"),
                user=user,
            )
            return self.json_response(result)
        except PublishError as ex:
            db.session.rollback()
            return json_error_response(str(ex), 400)
        except Exception as ex:  # pylint: disable=broad-except
            db.session.rollback()
            logger.exception("Failed to publish report")
            return json_error_response(utils.error_msg_from_exception(ex), 400)

    @expose("/api/<int:report_id>/unpublish/", methods=("POST",))
    @has_access_api
    def api_unpublish(self, report_id: int) -> FlaskResponse:
        """Detach a published report's chart/dashboard and clean up."""
        try:
            report = db.session.get(ReportDesigner, report_id)
            if report is None:
                return Response(status=404)
            return self.json_response(unpublish_report(report))
        except PublishError as ex:
            db.session.rollback()
            return json_error_response(str(ex), 400)
        except Exception as ex:  # pylint: disable=broad-except
            db.session.rollback()
            logger.exception("Failed to unpublish report")
            return json_error_response(utils.error_msg_from_exception(ex), 400)

    @expose("/api/<int:report_id>/export.csv/", methods=("GET",))
    @has_access_api
    def export_csv(self, report_id: int) -> FlaskResponse:
        report = db.session.get(ReportDesigner, report_id)
        if report is None:
            return Response(status=404)
        try:
            result = execute_report(report.get_definition())
            buffer = StringIO()
            writer = csv.writer(buffer)
            writer.writerow(result["columns"])
            for row in result["rows"]:
                writer.writerow([row.get(col, "") for col in result["columns"]])
            safe_name = SAFE_NAME_RE.sub("_", report.name) or "report"
            return Response(
                buffer.getvalue(),
                mimetype="text/csv",
                headers={
                    "Content-Disposition": (
                        f'attachment; filename="{safe_name}.csv"'
                    )
                },
            )
        except ReportDesignerError as ex:
            return json_error_response(str(ex), 400)
        except Exception as ex:  # pylint: disable=broad-except
            logger.exception("Failed to export report")
            return json_error_response(utils.error_msg_from_exception(ex), 400)

    @expose("/api/<int:report_id>/export.xlsx/", methods=("GET",))
    @has_access_api
    def export_xlsx(self, report_id: int) -> FlaskResponse:
        report = db.session.get(ReportDesigner, report_id)
        if report is None:
            return Response(status=404)
        try:
            result = execute_report(report.get_definition())
            buffer = _build_xlsx(result["columns"], result["rows"])
            safe_name = SAFE_NAME_RE.sub("_", report.name) or "report"
            return Response(
                buffer,
                mimetype=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                headers={
                    "Content-Disposition": (
                        f'attachment; filename="{safe_name}.xlsx"'
                    )
                },
            )
        except ReportDesignerError as ex:
            return json_error_response(str(ex), 400)
        except Exception as ex:  # pylint: disable=broad-except
            logger.exception("Failed to export report to Excel")
            return json_error_response(utils.error_msg_from_exception(ex), 400)

    @expose("/api/<int:report_id>/export.pdf/", methods=("GET",))
    @has_access_api
    def export_pdf(self, report_id: int) -> FlaskResponse:
        report = db.session.get(ReportDesigner, report_id)
        if report is None:
            return Response(status=404)
        try:
            definition = report.get_definition()
            result = execute_report(definition)
            from superset.views.report_designer.pdf_export import build_report_pdf

            data = build_report_pdf(
                report.name,
                report.description or "",
                definition,
                result,
                generated_by=utils.get_user_id() and str(utils.get_user_id()),
            )
            safe_name = SAFE_NAME_RE.sub("_", report.name) or "report"
            return Response(
                data,
                mimetype="application/pdf",
                headers={
                    "Content-Disposition": (
                        f'attachment; filename="{safe_name}.pdf"'
                    )
                },
            )
        except ReportDesignerError as ex:
            return json_error_response(str(ex), 400)
        except Exception as ex:  # pylint: disable=broad-except
            logger.exception("Failed to export report to PDF")
            return json_error_response(utils.error_msg_from_exception(ex), 400)

    @expose("/api/upload/", methods=("POST",))
    @has_access_api
    def api_upload(self) -> FlaskResponse:
        """Ingest an Excel/CSV file into a writable database as a dataset."""
        from superset.views.report_designer.excel_ingest import (
            ingest_excel,
            IngestError,
        )

        try:
            database_id = int(request.form.get("database_id") or 0)
            table_name = request.form.get("table_name") or None
            schema = request.form.get("schema") or None
            file_storage = request.files.get("file")
            if file_storage is None:
                return json_error_response(_("No file provided"), 400)
            database = db.session.get(Database, database_id) if database_id else None
            if database is None:
                return json_error_response(_("Database not found"), 400)
            payload = ingest_excel(
                database,
                file_storage.stream,
                file_storage.filename or "upload.xlsx",
                table_name=table_name,
                schema=schema,
            )
            return self.json_response(payload, 201)
        except (IngestError, ValueError) as ex:
            db.session.rollback()
            return json_error_response(str(ex), 400)
        except Exception as ex:  # pylint: disable=broad-except
            db.session.rollback()
            logger.exception("Failed to ingest spreadsheet")
            return json_error_response(utils.error_msg_from_exception(ex), 400)


def _build_xlsx(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    """Build an .xlsx payload with openpyxl (header styled)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    sheet.append(columns)
    for col_idx, _name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="102B6B", end_color="102B6B", fill_type="solid"
        )
    for row in rows:
        sheet.append([row.get(col, "") for col in columns])
    for idx, col in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        sheet.column_dimensions[letter].width = min(
            40, max(10, len(str(col)) + 2)
        )
    sheet.freeze_panes = "A2"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
